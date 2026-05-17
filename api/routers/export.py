"""
routers/export.py
Excel export endpoint for NGO analysts.
GET /api/export/excel
"""

from fastapi import APIRouter
from fastapi.responses import Response, JSONResponse
from api.database import get_pool
from datetime import datetime
import traceback
import io

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/export", tags=["Export"])

# ── Colour palette ────────────────────────────────────────────
NAVY        = "1B3A6B"
NAVY_LIGHT  = "2A4F8F"
GOLD        = "F5C842"
RED         = "B71C1C"
RED_LIGHT   = "FFEBEE"
ORANGE      = "E65100"
ORANGE_LIGHT= "FFF3E0"
AMBER       = "F9A825"
AMBER_LIGHT = "FFFDE7"
GREEN       = "2E7D32"
GREEN_LIGHT = "E8F5E9"
GREY_LIGHT  = "F8F9FA"
WHITE       = "FFFFFF"


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _border() -> Border:
    thin = Side(style="thin", color="E0E0E0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _risk_fill(score: int) -> PatternFill:
    if score >= 277: return _header_fill(RED_LIGHT)
    if score >= 199: return _header_fill(ORANGE_LIGHT)
    if score >= 126: return _header_fill(AMBER_LIGHT)
    if score >= 59:  return _header_fill("E3F2FD")
    return _header_fill(GREEN_LIGHT)


def _risk_label(score: int) -> str:
    if score >= 277: return "Critical"
    if score >= 199: return "High"
    if score >= 126: return "Moderate"
    if score >= 59:  return "Low"
    return "Stable"


def _severity_fill(severity: str) -> PatternFill:
    fills = {
        "Critical": _header_fill(RED_LIGHT),
        "Severe"  : _header_fill(ORANGE_LIGHT),
        "Moderate": _header_fill(AMBER_LIGHT),
    }
    return fills.get(severity, _header_fill(WHITE))


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, headers: list, row: int,
                      bg: str = NAVY, fg: str = "FFFFFF"):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill      = _header_fill(bg)
        cell.font      = _font(bold=True, color=fg, size=10)
        cell.alignment = _center()
        cell.border    = _border()


def _write_title(ws, title: str, subtitle: str,
                 ncols: int, generated_at: str):
    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill      = _header_fill(NAVY)
    c.font      = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=ncols)
    c2 = ws.cell(row=2, column=1,
                 value=f"{subtitle}  |  Generated: {generated_at}")
    c2.fill      = _header_fill(NAVY_LIGHT)
    c2.font      = _font(color="A8C4E8", size=9)
    c2.alignment = _center()
    ws.row_dimensions[2].height = 16

    # Blank spacer
    ws.row_dimensions[3].height = 6


# ── Sheet 1: District Risk Summary ────────────────────────────
def _build_district_sheet(ws, districts: list, generated_at: str):
    headers = [
        "#", "District", "Region", "Risk Level", "Risk Score",
        "Critical Spikes", "Severe Spikes", "Moderate Spikes",
        "Total Spikes", "Spike Rate %", "Avg Price (MWK)"
    ]
    _write_title(ws, "Malawi Food Security Monitor — District Risk Summary",
                 "WFP VAM Data · January 2020 – April 2026", len(headers), generated_at)
    _write_header_row(ws, headers, row=4)
    ws.freeze_panes = "A5"

    for i, d in enumerate(districts, 1):
        row = 4 + i
        risk_fill  = _risk_fill(d["risk_score"])
        risk_label = _risk_label(d["risk_score"])

        values = [
            i,
            d["name_1"],
            d["region"],
            risk_label,
            int(d["risk_score"]),
            d["critical_count"],
            d["severe_count"],
            d["moderate_count"],
            d["critical_count"] + d["severe_count"] + d["moderate_count"],
            round(float(d["spike_rate_pct"] or 0), 1),
            round(float(d["avg_price"] or 0), 0),
        ]

        for col, val in enumerate(values, 1):
            cell            = ws.cell(row=row, column=col, value=val)
            cell.border     = _border()
            cell.alignment  = _center() if col != 2 else _left()
            cell.font       = _font(size=10)

            # Color risk score and label columns
            if col in [4, 5]:
                cell.fill = risk_fill
                cell.font = _font(bold=True, size=10)

            # Alternate row shading
            elif i % 2 == 0:
                cell.fill = _header_fill(GREY_LIGHT)

    _set_col_widths(ws, [5, 18, 18, 14, 12, 14, 14, 14, 12, 12, 16])


# ── Sheet 2: Critical Spike Events ────────────────────────────
def _build_spikes_sheet(ws, spikes: list, generated_at: str):
    headers = [
        "Date", "District", "Region", "Market",
        "Commodity", "Price (MWK)", "% Change", "Z-Score", "Severity"
    ]
    _write_title(ws, "Malawi Food Security Monitor — Critical Spike Events",
                 "Price spikes ≥60% change + z-score ≥2.5", len(headers), generated_at)
    _write_header_row(ws, headers, row=4, bg=RED, fg="FFFFFF")
    ws.freeze_panes = "A5"

    for i, s in enumerate(spikes, 1):
        row  = 4 + i
        fill = _severity_fill(s["spike_severity"])

        values = [
            str(s["date"]),
            s["district"],
            s.get("region", ""),
            s["market"],
            s["commodity"],
            round(float(s["price"] or 0), 2),
            round(float(s["pct_change"] or 0), 1),
            round(float(s["zscore"] or 0), 2),
            s["spike_severity"],
        ]

        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = _border()
            cell.alignment = _center() if col not in [2, 3, 4, 5] else _left()
            cell.font      = _font(size=10)
            if col == 9:  # severity column
                cell.fill = fill
                cell.font = _font(bold=True, size=10)
            elif i % 2 == 0:
                cell.fill = _header_fill(GREY_LIGHT)

    _set_col_widths(ws, [14, 16, 18, 20, 20, 14, 12, 10, 12])


# ── Sheet 3: Commodity Analysis ───────────────────────────────
def _build_commodity_sheet(ws, commodities: list, generated_at: str):
    headers = [
        "Commodity", "Critical Events", "Severe Events",
        "Moderate Events", "Total Spikes", "Spike Rate %"
    ]
    _write_title(ws, "Malawi Food Security Monitor — Commodity Analysis",
                 "Price shock frequency by commodity", len(headers), generated_at)
    _write_header_row(ws, headers, row=4, bg=NAVY_LIGHT, fg="FFFFFF")
    ws.freeze_panes = "A5"

    for i, c in enumerate(commodities, 1):
        row = 4 + i
        values = [
            c["commodity"],
            c["critical_count"],
            c["severe_count"],
            c["moderate_count"],
            c["total_spikes"],
            round(float(c["spike_rate"] or 0), 1),
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = _border()
            cell.alignment = _left() if col == 1 else _center()
            cell.font      = _font(size=10)
            if col == 2 and int(c["critical_count"]) > 0:
                cell.fill = _header_fill(RED_LIGHT)
                cell.font = _font(bold=True, size=10)
            elif i % 2 == 0:
                cell.fill = _header_fill(GREY_LIGHT)

    _set_col_widths(ws, [25, 16, 14, 16, 14, 14])


# ── Sheet 4: Monitoring Gaps ──────────────────────────────────
def _build_gaps_sheet(ws, gaps: list, generated_at: str):
    headers = [
        "District", "Region", "Risk Score",
        "Markets Monitored", "Monitoring Status"
    ]
    _write_title(ws, "Malawi Food Security Monitor — Market Monitoring Gaps",
                 "Districts with high risk but insufficient market coverage",
                 len(headers), generated_at)
    _write_header_row(ws, headers, row=4, bg=ORANGE, fg="FFFFFF")
    ws.freeze_panes = "A5"

    status_fills = {
        "CRITICAL GAP"  : _header_fill(RED_LIGHT),
        "HIGH GAP"      : _header_fill(ORANGE_LIGHT),
        "MODERATE GAP"  : _header_fill(AMBER_LIGHT),
        "ADEQUATE"      : _header_fill(GREEN_LIGHT),
        "NO MONITORING" : _header_fill("EEEEEE"),
    }

    for i, g in enumerate(gaps, 1):
        row    = 4 + i
        status = g["monitoring_status"]
        values = [
            g["district"],
            g["region"],
            int(g["risk_score"]),
            int(g["market_count"]),
            status,
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = _border()
            cell.alignment = _left() if col in [1, 2, 5] else _center()
            cell.font      = _font(size=10)
            if col == 5:
                cell.fill = status_fills.get(status, _header_fill(WHITE))
                cell.font = _font(bold=True, size=10)
            elif i % 2 == 0:
                cell.fill = _header_fill(GREY_LIGHT)

    _set_col_widths(ws, [20, 20, 14, 18, 20])


# ── Endpoint ──────────────────────────────────────────────────
@router.get("/excel")
async def export_excel():
    """
    Generate and download a formatted Excel workbook
    with 4 sheets of food security data for NGO analysts.
    """
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:

            districts = await conn.fetch("""
                SELECT name_1, region, risk_score, critical_count,
                       severe_count, moderate_count, spike_rate_pct, avg_price
                FROM districts_risk
                ORDER BY risk_score DESC
            """)

            spikes = await conn.fetch("""
                SELECT date::text, district, region, market, commodity,
                       price, pct_change, zscore, spike_severity
                FROM spikes
                ORDER BY date DESC, spike_severity
            """)

            commodities = await conn.fetch("""
                SELECT
                    commodity,
                    SUM(CASE WHEN spike_severity='Critical' THEN 1 ELSE 0 END) AS critical_count,
                    SUM(CASE WHEN spike_severity='Severe'   THEN 1 ELSE 0 END) AS severe_count,
                    SUM(CASE WHEN spike_severity='Moderate' THEN 1 ELSE 0 END) AS moderate_count,
                    COUNT(*) AS total_spikes,
                    ROUND(COUNT(*)::numeric / NULLIF(
                        (SELECT COUNT(*) FROM spikes), 0
                    ) * 100, 2) AS spike_rate
                FROM spikes
                GROUP BY commodity
                ORDER BY critical_count DESC
            """)

            gaps = await conn.fetch("""
                SELECT
                    d.name_1 AS district, d.region, d.risk_score,
                    COUNT(m.ogc_fid) AS market_count,
                    CASE
                        WHEN COUNT(m.ogc_fid) = 0                         THEN 'NO MONITORING'
                        WHEN d.risk_score > 400                           THEN 'CRITICAL GAP'
                        WHEN d.risk_score > 200 AND COUNT(m.ogc_fid) < 8 THEN 'CRITICAL GAP'
                        WHEN d.risk_score > 100 AND COUNT(m.ogc_fid) < 4 THEN 'HIGH GAP'
                        WHEN d.risk_score > 50  AND COUNT(m.ogc_fid) < 2 THEN 'MODERATE GAP'
                        ELSE 'ADEQUATE'
                    END AS monitoring_status
                FROM districts_risk d
                LEFT JOIN markets m ON ST_Within(m.wkb_geometry, d.wkb_geometry)
                GROUP BY d.name_1, d.region, d.risk_score
                ORDER BY d.risk_score DESC
            """)

        generated_at = datetime.now().strftime("%d %B %Y %H:%M")

        # Build workbook
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        ws1 = wb.create_sheet("District Risk Summary")
        ws2 = wb.create_sheet("Critical Spike Events")
        ws3 = wb.create_sheet("Commodity Analysis")
        ws4 = wb.create_sheet("Monitoring Gaps")

        # Tab colors
        ws1.sheet_properties.tabColor = NAVY
        ws2.sheet_properties.tabColor = RED
        ws3.sheet_properties.tabColor = NAVY_LIGHT
        ws4.sheet_properties.tabColor = ORANGE

        _build_district_sheet(ws1,   [dict(r) for r in districts],   generated_at)
        _build_spikes_sheet(ws2,     [dict(r) for r in spikes],       generated_at)
        _build_commodity_sheet(ws3,  [dict(r) for r in commodities],  generated_at)
        _build_gaps_sheet(ws4,       [dict(r) for r in gaps],         generated_at)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"malawi_food_security_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return Response(
            content      = buffer.getvalue(),
            media_type   = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers      = {"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        return JSONResponse(status_code=500, content={
            "error": str(e),
            "trace": traceback.format_exc(),
        })